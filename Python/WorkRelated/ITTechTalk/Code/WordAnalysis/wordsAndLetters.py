import os
import re
import string

import openpyxl


"""
TODO: Implement a feature so that each book is given its own sheet for analysis
"""
path_to_texts = (
    "C:/dev/workspace-vscode/Python/WorkRelated/ITTechTalk/Code/WordAnalysis/texts"
)

ascii_con = {}
a = 65
for i in range(26):
    ascii_con[a] = a + 32
    a += 1


def find_all_non_numslets():
    paths = get_paths(path_to_texts)
    all_used_characters = set()
    for path in paths:
        with open(path, encoding="utf-8") as f:
            content = f.read()
            all_used_characters.update(set(content))
    all_used_list = [i for i in all_used_characters]
    special_characters = []
    for x in all_used_list:
        if str(x).lower() not in "abcdefghijklmnopqrstuvwxyz":
            special_characters.append(x)
    u = ""
    special_characters.remove("\n")
    for i in special_characters:
        u += i
    return u


def ascii_trans(word):
    new_word = ""
    for let in word:
        if ord(let) not in ascii_con.keys():
            new_word += let
        else:
            new_word += chr(ascii_con[ord(let)])
    return new_word


def word_count():
    paths = get_paths(path_to_texts)
    special_characters = find_all_non_numslets()
    regex = re.compile("[{}0-9]".format(re.escape(string.punctuation)))
    words = {}
    for path in paths:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                for word in line.split():
                    # Take out numbers and punctuation
                    new_word = re.sub(
                        regex, "", word
                    )  
                    # Take out goofy letters
                    new_word = new_word.translate(
                        str.maketrans("", "", special_characters)
                    )  
                    # Normalize ascii values
                    new_word = ascii_trans(new_word)
                    if new_word not in words:
                        words[new_word] = 1
                    else:
                        words[new_word] += 1
    f.close()
    return words


def generate_word_file(wawb, words):
    wordAnalysis = wawb["Word Analysis"]
    counter = 4
    total_words = 0
    for y in words:
        total_words += words[y]
    for x in words:
        wordAnalysis.cell(row=counter, column=1).value = str(x).lower()
        wordAnalysis.cell(row=counter, column=2).value = str(
            [(str(ord(i)) + " ") for i in x]
        )
        wordAnalysis.cell(row=counter, column=3).value = words[x]
        wordAnalysis.cell(row=counter, column=4).value = len(x)
        wordAnalysis.cell(row=counter, column=5).value = words[x] / total_words
        counter += 1
    wordAnalysis.cell(row=1, column=2).value = total_words
    wawb.save("Analysis.xlsx")


def letter_count():
    paths = get_paths(path_to_texts)
    letters = {}
    for path in paths:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                for char in line:
                    if char.isalpha():
                        lower = char.lower()
                        if lower not in letters:
                            letters[lower] = 1
                        else:
                            letters[lower] += 1
    f.close()
    return letters


def generate_letter_file(wawb, letters):
    letterAnalysis = wawb["Letter Analysis"]
    total = 0
    counter = 4
    for x in letters:
        if x in "abcdefghijklmnopqrstuvwxyz":
            total += letters[x]
            letterAnalysis.cell(row=counter, column=1).value = x
            letterAnalysis.cell(row=counter, column=2).value = letters[x]
            counter += 1
    letterAnalysis.cell(row=1, column=2).value = total
    wawb.save("Analysis.xlsx")


def setup_wb(wawb):
    wawb.create_sheet(index=0, title="Word Analysis")
    wawb.create_sheet(index=1, title="Letter Analysis")

    # Creating Words sheet
    wordAnalysis = wawb["Word Analysis"]
    wordAnalysis["A1"] = "Total:"
    wordAnalysis["A3"] = "Words"
    wordAnalysis["B3"] = "ASCII"
    wordAnalysis["C3"] = "Occurrences"
    wordAnalysis["D3"] = "Word Length"
    wordAnalysis["E3"] = "Ratio"

    # Creating Letters sheet
    letterAnalysis = wawb["Letter Analysis"]
    letterAnalysis["A1"] = "Total:"
    letterAnalysis["A3"] = "Letters"
    letterAnalysis["B3"] = "Occurrences"
    letterAnalysis["C3"] = "Ratio"
    wawb.save("Analysis.xlsx")


def get_paths(main_path):
    text_file_names = os.listdir(main_path)
    return [main_path + "/" + c for c in text_file_names]



def main():
    # list_of_texts = get_paths(path_to_texts)
    wb = openpyxl.Workbook()
    setup_wb(wb)
    words = word_count()
    letters = letter_count()
    generate_word_file(wb, words)
    generate_letter_file(wb, letters)


if __name__ == "__main__":
    main()
