import re
import os
import openpyxl
import string

"""
TODO: Implement a feature so that each book is given its own sheet for analysis
"""

def word_analysis(paths, wawb):
    regex = re.compile('[%s]' % re.escape(string.punctuation))
    wordAnalysis = wawb['Word Analysis']
    words = {}
    for path in paths:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                for word in line.split():
                    new_word = regex.sub('', word)
                    if new_word not in words:
                        words[new_word] = 1
                    else:
                        words[new_word] += 1
    f.close()
    counter = 2
    total_words = 0
    for x in words:
        wordAnalysis.cell(row=counter, column=1).value = x
        wordAnalysis.cell(row=counter, column=2).value = words[x]
        wordAnalysis.cell(row=counter, column=3).value = len(x)
        total_words += words[x]
        counter += 1
    wordAnalysis.cell(row=2, column=4).value = total_words
    wawb.save("Analysis.xlsx")


def letter_analysis(paths, wawb):
    letterAnalysis = wawb['Letter Analysis']
    letters = {}
    for path in paths:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                for char in line:
                    if char.isalpha():
                        lower = char.lower()
                        if lower not in letters:
                            letters[lower] = 1
                        else:
                            letters[lower] += 1
    f.close()
    counter = 2
    letterAnalysis['A1'] = 'Letter'
    letterAnalysis['B1'] = 'Occurrences'
    for x in letters:
        letterAnalysis.cell(row=counter, column=1).value = x
        letterAnalysis.cell(row=counter, column=2).value = letters[x]
        counter += 1
    wawb.save("Analysis.xlsx")


def setup_wb(wawb):
    wawb.create_sheet(index=0, title='Word Analysis')
    wawb.create_sheet(index=1, title='Letter Analysis')
    wordAnalysis = wawb['Word Analysis']
    wordAnalysis['A1'] = 'Word'
    wordAnalysis['B1'] = 'Occurrences'
    wordAnalysis['C1'] = 'Word Length'
    wordAnalysis['D1'] = 'Total Words'
    letterAnalysis = wawb['Letter Analysis']
    letterAnalysis['A1'] = 'Letter'
    letterAnalysis['B1'] = 'Occurrences'
    wawb.save("Analysis.xlsx")


def main():
    path_to_texts = "C:/dev/workspace-vscode/Python/WorkRelated/ITTechTalk/Code/Word_Analysis/texts"
    text_file_names = os.listdir(path_to_texts)
    list_of_texts = [path_to_texts + '/' + c for c in text_file_names]
    wb = openpyxl.Workbook()
    setup_wb(wb)
    word_analysis(list_of_texts, wb)
    letter_analysis(list_of_texts, wb)


if __name__ == '__main__':
    main()
