import math
import os
import re
import string
from string import punctuation

import openpyxl

from Code.WordAnalysis.wordsAndLetters import ascii_trans, word_count

path_to_passwords = "C:/dev/workspace-vscode/Python/WorkRelated/ITTechTalk/Code/PasswordStrengthCalculators/passwords"
files = os.listdir(path_to_passwords)
path_to_files = [path_to_passwords + "/" + x for x in files]

path_to_wb = "C:/dev/workspace-vscode/Python/WorkRelated/ITTechTalk/Analysis.xlsx"


def num_format(num):
    return f"{num:,d}"


def check_if_pw_exists(pw):
    for path in path_to_files:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line == pw:
                    return "Exists"


def check_for_words(pw):
    list_of_words = word_count()
    pw = str(pw).lower()
    matches = []
    for word, value in list_of_words.items():
        word = str(word).lower()
        if word in pw:
            match_start = pw.find(word)
            if len(word) >= 3:
                if word == pw[match_start : match_start + len(word)]:
                    matches.append([word, value])
    return matches


def check_known(pw):
    pw = str(pw)
    for path in path_to_files:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if pw.lower() == line.lower():
                    return line
    return None


def check_consecutive(pw):
    last = ""
    results = []
    for letter in pw:
        if letter == last:
            results[-1] = (letter, results[-1][1] + 1)
        else:
            results.append((letter, 1))
            last = letter
    return results


def uPass(pw):
    l = len(pw)
    e = 0
    nones = [
        (re.search(r"[a-z]", pw), 26),
        (re.search(r"[A-Z]", pw), 26),
        (re.search(r"[0-9]", pw), 10),
        (re.search(r"[%s]" % re.escape(string.punctuation), pw), 33),
    ]
    counter = 0
    for i, k in nones:
        if i != None:
            e += k
        else:
            counter += 1
    if counter == 0:
        return [e, math.pow(e, l)]
    else:
        return [e, math.pow(e, l) / math.pow(counter, 2)]


def detect_missing(pw):
    nones = [
        (re.search(r"[a-z]", pw)),
        (re.search(r"[A-Z]", pw)),
        (re.search(r"[0-9]", pw)),
        (re.search(r"[%s]" % re.escape(string.punctuation), pw)),
    ]
    statement = {
        1: "Lowercase letters",
        2: "Uppercase letters",
        3: "Numbers",
        4: "Symbols",
    }
    values = {1: 26, 2: 26, 3: 10, 4: 33}
    missing = []
    for i in range(len(nones)):
        if nones[i] == None:
            missing.append([statement[i + 1], values[i + 1]])
    return missing


def get_rate(words):
    wb = openpyxl.load_workbook(path_to_wb)
    wa = wb["Word Analysis"]
    ws = {}
    for row in range(4, wa.max_row + 1):
        word = wa["A" + str(row)].value
        occurence = wa["C" + str(row)].value
        ws[word] = occurence
    totals = []
    for i, k in words:
        if i in ws:
            totals.append([i, ws[i]])
    return totals


def main():
    pw = str(input("Enter a FAKE password: "))
    basic_check = uPass(pw)
    missing = detect_missing(pw)
    matches = check_for_words(pw)
    cons = check_consecutive(pw)
    known_pw = check_known(pw)
    occurances_rate = get_rate(matches)
    
    a = basic_check[1]
    raw = int(math.pow(basic_check[0], len(pw)))

    # Printing Facts: Length and Entropy
    print(
        "\tYour password facts:\n\t - PW Length = {}\n\t - PW Entropy = {}".format(
            len(pw), basic_check[0]
        )
    )
    # Printing missing values
    if len(missing) > 0:
        for i, k in missing:
            print("\t\t - You are missing {} which are worth {}".format(i, k))
    # Printing Strength/Bytes
    print(
        "\t - Raw Strength = {}\n\t\t - or {} bytes".format(
            raw, math.log2(math.pow(int(basic_check[0]), len(pw)))
        )
    )

    if len(missing) > 0:
        print(
            "\nSince passwords become exponentially easier to crack the shorter they are and the less entropy they have, we will take"
        )   
        print(
            "your Raw Strength and divide it by the number of missing sets of characters (if any) to the power of 2 ({}^2, in this case)".format(
                len(missing)
            )
        )
        print("\t{} / {}^2 = {}".format(raw, len(missing), a))

    print("\nStarting Strength: " + str(int(a)))

    # Changing A based on length of password
    if len(pw) < 6:
        a = a / 10
        print("Length Mod: " + str(int(a)) + "\n\t^L = " + str(len(pw)))
    elif len(pw) < 12:
        a = a / 5
        print("Length Mod: " + str(int(a)) + "\n\t^L = " + str(len(pw)))
    elif len(pw) < 15:
        a = a / 2
        print("Length Mod: " + str(int(a)) + "\n\t^L = " + str(len(pw)))
    else:
        print("Length Mod (none): " + str(int(a)) + "\n\t^L = " + str(len(pw)))

    # Changing A based on matches to dictionary words found in the password itself
    if len(matches) > 0:
        a = a / (len(matches) / 0.25)
        print(
            "Checking for words in your password: "
            + str(int(a))
            + "\n\t^"
            + str(matches)
            + "^"
        )
        occurances = 0
        for i, j in occurances_rate:
            occurances += j
        occurances = occurances / len(matches)
        a = a / occurances
        print("Checking for consecutive letters: " + str(int(a)) + "\n\t^" + str(cons) + "^")

    # Checking for consecutive letters
    cons_total = 0
    for x, y in cons:
        if y > 1:
            cons_total += 1
    if cons_total > 1:
        a = a / cons_total
        print(
            "You used "
            + str(cons_total)
            + " different consecutive letters:\n\t"
            + str(int(a))
        )

    # Checking if known password
    if known_pw != None:
        a = math.sqrt(a)
        print("Your password is a known password: " + str(int(a)))

    # End strength
    print("Ending Strength: " + str(int(a)))
    print("\tOr " + str(int(math.log2(a))) + " bytes")


if __name__ == "__main__":
    main()
